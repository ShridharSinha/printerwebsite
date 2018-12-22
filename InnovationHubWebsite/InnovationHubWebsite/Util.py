from datetime import *
from .models  import *
from random import randint
from openpyxl import load_workbook
from openpyxl import Workbook
from django.core.files.storage import default_storage
#import Blender
import os
import sys
#import bpy
#from pymesh import stl, obj

class Util:
    def handle_file(self, f, name, user, jobid):

        date = '_' + str(datetime.now().day) + '_' + str(datetime.now().month) + '_' + str(datetime.now().year) + '_' + str(datetime.now().hour)+ '_' + str(datetime.now().minute)+ '_' + str(datetime.now().second) + '_'

        name += date + user.first_name + '_' + user.last_name
        name1 = str(jobid) + '_' + name + '.stl'

        #path = 'UploadedFiles/' + name
        #path = 'The.stl'

        #open(path, 'a').close()

        #with open(path, 'wb+') as destination:
        #    for chunk in f.chunks():
        #        destination.write(chunk)

        file = default_storage.open(name1, 'w')
        file.write(f)
        file.close()

        #path = file.get()['Body'].read()
        path = file.name
        #print("Path:" + path)

        path1 = path
        #path2 = 'static/JS/3DModels/' + name + '.obj'

        #f2 = self.convertSTLtoOBJ(path, path2)
        #f2 = self.convertSTLtoOBJ(path)
        #f2 = self.saveSTLasOBJ(path1, path2, f)
        f2 = f

        name2 = str(jobid) + '_' + name + '.obj'
        file2 = default_storage.open(name2, 'w')

        path2 = file2.name

        #open(path2, 'a').close()

        #with open(path2, 'wb+') as destination:
        #    for chunk in f2.chunks():
        #        destination.write(chunk)

        #return(path1, path2)
        return(path1, path2)

    def convertSTLtoOBJ(self, f):
        os.system("cd static\\JS\\3DModels && blender -b Empty.obj.blend -P blender.py "+f+".stl")
        k=sys.argv[1]
        #print(k)
        #path = 'static/JS/3DModels'
        return(open(f2, 'a'))
        #blender = Blender()
        #blender.import_export(f)



    #def saveSTLasOBJ(self, path1, path2, f):
    #    PyMesh = pymesh()
    #    mesh = PyMesh.load_mesh(path1)
    #    PyMesh.save_mesh(path2, mesh)
    #def saveSTLasOBJ(self, path1, path2, f):
        #mesh = pymesh.meshio.load_mesh(path1, drop_zero_dim=False)
        #pymesh.meshio.save_mesh(path2, mesh)
        #pymesh.test()
        #m = stl.Stl(path1)
        #m.save_obj(path2)


    def getPrintStartTime(self):
        return(datetime.now())

    def getPrintEndTime(self, f):
        total_wait_time = 0
        total_print_num = 0
        stats = list(Statistic.objects.all())

        for stat in stats:
            total_wait_time = total_wait_time + stat.wait_time
            total_print_num = total_print_num + stat.print_num

        average_wait_time = 0
        if(not(total_print_num == 0)):
            average_wait_time = total_wait_time/total_print_num

        day_num = (average_wait_time/86400) + 4

        return(datetime.now() + timedelta(days=day_num) )

    def getCurrentYear(self):
        return(datetime.now().strftime("%y"))


    def getCurrentMonth(self):
        month_num  = datetime.now().strftime("%m")
        month_name = self.convertMonthNumToName(month_num)

        #print(month_name)
        return(month_name)

    def convertMonthNumToName(self, num):
        if(num == 1 or num == '01'):
            return('January')
        elif(num == 2 or num == '02'):
            return('February')
        elif(num == 3 or num == '03'):
            return('March')
        elif(num == 4 or num == '04'):
            return('April')
        elif(num == 5 or num == '05'):
            return('May')
        elif(num == 6 or num == '06'):
            return('June')
        elif(num == 7 or num == '07'):
            return('July')
        elif(num == 8 or num == '08'):
            return('August')
        elif(num == 9 or num == '09'):
            return('September')
        elif(num == 10 or num == '10'):
            return('October')
        elif(num == 11 or num == '11'):
            return('November')
        else:
            return('December')

    def clearStatistics(self):
        months = list(Statistic.objects.all())

        for month in months:
            month.clear()
            month.save()


    def getPrinterName(self):
        #printer = ['Thor', 'Artemis', 'Zeus']
        #return(printer[randint(0, 2)])
        Thor    = list(Job.objects.filter(printer_name = 'Thor').filter(status = 'in Queue'))
        Zeus    = list(Job.objects.filter(printer_name = 'Zeus').filter(status = 'in Queue'))
        Artemis = list(Job.objects.filter(printer_name = 'Artemis').filter(status = 'in Queue'))

        prints = [len(Thor), len(Zeus), len(Artemis)]

        printer = ''

        if(prints[0] < prints[1] and prints[0] < prints[2]):
            printer = 'Thor'
        elif(prints[1] < prints[2]):
            printer = 'Zeus'
        else:
            printer = 'Artemis'

        return(printer)



    def getProfile(self, user):
        profiles = list(Profile.objects.all())

        for i in range(0, len(profiles)):
            if(profiles[i].equals(user)):
                return profiles[i]


    def getQuota(self, user):
        if(user.is_authenticated):
            profiles = list(Profile.objects.all())

            for i in range(0, len(list(profiles))):
                if(profiles[i].equals(user)):
                    profile = profiles[i]

            timeS = profile.quota

            timeH = str(int(timeS / 3600))
            timeS = timeS % 3600
            timeM = str(int(timeS / 60))
            timeS = str(timeS % 60)

            if(len(timeH) < 2):
                timeH = '0' + timeH
            if(len(timeM) < 2):
                timeM = '0' + timeM
            if(len(timeS) < 2):
                timeS = '0' + timeS

            quota = timeH + ':' + timeM + ':' + timeS

            #context = {'Quota' : '00:31:23'}
            return({'Quota' : quota})
        else:
            return({'Quota' : '--:--:--'})

    def changeGrade(self, diff):
        profiles = list(Profile.objects.all())

        for p in profiles:
            p.grade += diff
            p.save()


    def getActiveUserNum(self):
        profiles = list(Profile.objects.all())
        jobs     = list(Job.objects.all())
        num = 0
        #print(len(profiles))

        for i in range(len(profiles) - 1, -1, -1):
            #print(i)
            if(profiles[i].user.is_superuser):
                profiles.pop(i)
                #print('pop')

        for i in range(0, len(profiles)):
            for j in range(0, len(jobs)):
                if(jobs[j].fk_profile == profiles[i] and self.convertMonthNumToName(jobs[j].upload_time.strftime("%m")) == self.getCurrentMonth() and jobs[j].upload_time.strftime("%y") == self.getCurrentYear()):
                    num = num+1
                    break

        print(num)

        return(num)


    def getVeryActiveUserNum(self):
        profiles = list(Profile.objects.all())

        for i in range(len(profiles) - 1, -1, -1):
            if(not(profiles[i].quota == 0 and not(profiles[i].user.is_superuser))):
                profiles.pop(i)
        print(len(profiles))
        return(len(profiles))


    """def readFrom(a,b):
          wb = load_workbook(filename=a)
          ws = wb[b]
          data=[]
          for i in range(1,len(list(ws.rows))):
             row=list(ws.rows)[i]
             k=[]
             for cell in row:
                k.append(cell.value)
             data.append(k)
          return(data)

    def writeTo(a,b,c):
       wb = load_workbook(a)
       ws = wb[b]
       for i in range(1,len(list(ws.rows))):
             row=list(ws.rows)[i]
             count=0
             print(len(list(row)))
             for cell in row:
                try:
                   cell.value=c[i-1][count]
                except:
                   cell.value=None
                count=count+1
       try:
          for j in range(i,len(c)):
             ws.append(c[j-1])
       except:
          print('')
       wb.save(a)"""


    def readFrom(self, file_name, sheet_name):
          wb = load_workbook(filename=file_name)
          ws = wb[sheet_name]
          data=[]
          for i in range(0,len(list(ws.rows))):
             row=list(ws.rows)[i]
             k=[]
             for cell in row:
                k.append(cell.value)
             data.append(k)
          return(data)

    def writeTo(self, file_name, file_data):
       wb = Workbook()
       ws = wb.active

       #file_name = 'InnovationHubWebsite/' + file_name

       for i in range(0,len(file_data)):
           ws.append(file_data[i])
           #print(file_data[i])
           #print('appending')
       wb.save(file_name)
